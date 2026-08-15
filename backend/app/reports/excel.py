from __future__ import annotations
from io import BytesIO
from openpyxl import Workbook
from .schemas import Report

class ExcelReportGenerator:
    def generate(self, report: Report) -> bytes:
        workbook = Workbook(); overview = workbook.active; overview.title = "Overview"
        overview.append(["Report", report.title]); overview.append(["Executive Summary", report.executive_summary])
        for key, value in report.dataset_overview.items(): overview.append([key, str(value)])
        kpis = workbook.create_sheet("KPIs"); kpis.append(["KPI", "Value", "Operation"])
        for item in report.kpis: kpis.append([item["name"], item["value"], item["operation"]])
        quality = workbook.create_sheet("Data Quality"); quality.append(["Metric", "Value"])
        for key, value in report.data_quality.items(): quality.append([key, str(value)])
        insights = workbook.create_sheet("Insights"); insights.append(["Title", "Description", "Recommendation"])
        for item in report.insights: insights.append([item.get("title"), item.get("description"), item.get("recommendation")])
        if report.forecast:
            sheet = workbook.create_sheet("Forecast"); sheet.append(["Date", "Prediction", "Lower", "Upper"])
            for point in report.forecast.get("forecast", []): sheet.append([point.get("date"), point.get("prediction"), point.get("lower"), point.get("upper")])
        buffer = BytesIO(); workbook.save(buffer); return buffer.getvalue()
